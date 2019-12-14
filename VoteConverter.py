#!/usr/bin/env python
# coding: utf-8

# In[57]:


import openpyxl
import sys

print(sys.argv)

path = "test.xlsx"

if len(sys.argv) < 2:
    print("No file defined: test.xlsx is beeing used")
    exit()

# In[58]:


wb = openpyxl.load_workbook(path)
sheet = wb.active

# In[62]:


out = open("vote.csv", "w")
print(sheet.max_row-1, file=out)
for zettel in range(2,sheet.max_column):
    vote = [None for x in range(2,sheet.max_row+1)]
    for kandidat in range(2,sheet.max_row+1):
        #print(sheet[zettel][kandidat].value, end = "\t")
        #print(sheet[kandidat][zettel].value)
        if sheet[kandidat][zettel].value:
            vote[sheet[kandidat][zettel].value -1] = sheet[kandidat][1].value
            vote[sheet[kandidat][zettel].value -1] = sheet[kandidat][0].value #id
    vote = [str(v) for v in vote if v]
    reihung = " ".join(vote)
    print(1, reihung, 0, file=out)
out.close()


# In[ ]:




